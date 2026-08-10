import io
import re

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# CONFIGURATION
# ============================================================

MAIN_RACES = [
    "MELAYU",
    "CINA",
    "INDIA",
    "LAIN-LAIN",
]

PARTY_COLS = [
    "UMNO",
    "PKR",
    "PAS",
    "PPBM",
]

SIKAP_COLS = [
    "PUTIH",
    "KELABU",
    "HITAM",
]

DEFAULT_AGE_GROUPS = [
    "18-21",
    "22-30",
    "31-40",
    "41-50",
    "51-60",
    "61+",
]


# ============================================================
# COLUMN HELPERS
# ============================================================

def get_col(df, possible_names):
    col_map = {
        str(column).lower().strip(): column
        for column in df.columns
    }

    for name in possible_names:
        key = str(name).lower().strip()

        if key in col_map:
            return col_map[key]

    return None


def clean_service_no(value):
    if pd.isna(value):
        return ""

    value = str(value).strip().upper()

    if value in {"", "NAN", "NONE", "NULL"}:
        return ""

    return value


def clean_filename(value):
    if pd.isna(value):
        return "OUTPUT"

    name = str(value).strip().upper()

    name = re.sub(
        r'[\\/:*?"<>|]',
        " ",
        name,
    )

    name = " ".join(name.split())

    return name if name else "OUTPUT"


def clean_sheet_name(value, existing_names):
    if pd.isna(value):
        name = "DUN"
    else:
        name = str(value).strip().upper()

    name = re.sub(
        r'[\\/?*\[\]:]',
        " ",
        name,
    )

    name = " ".join(name.split())

    if not name:
        name = "DUN"

    name = name[:31]

    base = name
    counter = 2

    while name in existing_names:
        suffix = f" ({counter})"

        name = (
            base[:31 - len(suffix)]
            + suffix
        )

        counter += 1

    existing_names.add(name)

    return name


# ============================================================
# DUN / DM HELPERS
# ============================================================

def kod_dun_digits(kod_dun):
    if pd.isna(kod_dun):
        return ""

    value = str(kod_dun).strip()

    value = value.split(".")[0]

    return re.sub(
        r"\D",
        "",
        value,
    )


def kod_dun_sort_key(kod_dun):
    digits = kod_dun_digits(kod_dun)

    if not digits:
        return -1

    return int(digits)


def format_sheet_label(kod_dun, nama_dun):
    digits = kod_dun_digits(kod_dun)

    last2 = (
        digits[-2:].zfill(2)
        if digits
        else "00"
    )

    nama = (
        ""
        if pd.isna(nama_dun)
        else str(nama_dun).strip()
    )

    return f"N.{last2} {nama}"


def format_kod_dm(value):
    if pd.isna(value):
        return ""

    kod = str(value).strip()

    if kod in {
        "",
        "None",
        "none",
        "nan",
        "NaN",
    }:
        return ""

    kod = kod.split(".")[0]

    kod = kod.zfill(7)

    return (
        f"{kod[:3]}/"
        f"{kod[3:5]}/"
        f"{kod[5:]}"
    )


# ============================================================
# RACE / SIKAP / SERVICE CLASSIFICATION
# ============================================================

def normalise_race(value):
    if pd.isna(value):
        return "LAIN-LAIN"

    race = str(value).strip().upper()

    if race in {
        "MELAYU",
        "CINA",
        "INDIA",
    }:
        return race

    return "LAIN-LAIN"


def normalise_sikap(value):
    if pd.isna(value):
        return ""

    sikap = str(value).strip().upper()

    if sikap in {
        "KELABU-LAMA",
        "KELABU-BARU",
    }:
        return "KELABU"

    if sikap in {
        "PUTIH",
        "KELABU",
        "HITAM",
    }:
        return sikap

    return ""


def classify_awal(value):
    value = clean_service_no(value)

    if value == "":
        return ""

    if (
        value.startswith("G")
        or value.startswith("RF")
    ):
        return "POLIS"

    if value.startswith("T"):
        return "ASKAR"

    return "PENGUNDI AWAL"


def is_polis(value):
    value = clean_service_no(value)

    return (
        value.startswith("G")
        or value.startswith("RF")
    )


def is_askar(value):
    value = clean_service_no(value)

    return value.startswith("T")


# ============================================================
# AGE GROUPS
# ============================================================

def parse_age_groups(age_groups):
    if age_groups is None:
        age_groups = DEFAULT_AGE_GROUPS

    if len(age_groups) != 6:
        raise ValueError(
            "Exactly 6 age groups are required."
        )

    cleaned = [
        str(value).strip()
        for value in age_groups
    ]

    if any(
        not value
        for value in cleaned
    ):
        raise ValueError(
            "All 6 age groups must be filled in."
        )

    if len(set(cleaned)) != 6:
        raise ValueError(
            "Age groups must be unique."
        )

    parsed = []

    for index, label in enumerate(cleaned):

        if (
            index == 5
            and label.endswith("+")
        ):
            start_text = label[:-1].strip()

            if not start_text.isdigit():
                raise ValueError(
                    f"Invalid age group: {label}"
                )

            parsed.append({
                "label": label,
                "min": int(start_text),
                "max": None,
            })

            continue

        match = re.fullmatch(
            r"(\d+)\s*-\s*(\d+)",
            label,
        )

        if not match:
            raise ValueError(
                f"Invalid age group format: {label}. "
                "Use formats such as 18-21 or 61+."
            )

        minimum = int(match.group(1))
        maximum = int(match.group(2))

        if minimum > maximum:
            raise ValueError(
                f"Invalid age range: {label}"
            )

        parsed.append({
            "label": label,
            "min": minimum,
            "max": maximum,
        })

    return cleaned, parsed


def get_age_group(
    value,
    parsed_age_groups,
):
    try:
        if pd.isna(value):
            return ""

        age = int(float(value))

    except (
        ValueError,
        TypeError,
    ):
        return ""

    for group in parsed_age_groups:

        minimum = group["min"]
        maximum = group["max"]

        if maximum is None:
            if age >= minimum:
                return group["label"]

        elif (
            minimum <= age <= maximum
        ):
            return group["label"]

    return ""


# ============================================================
# PERCENTAGE
# ============================================================

def pct(part, total):
    if not total:
        return 0

    return round(
        part / total * 100,
        1,
    )


# ============================================================
# HEADERS
# ============================================================

def build_headers(age_groups):

    headers = [
        "KOD DM",
        "NAMA DM",
        "JUMLAH",

        "LELAKI",
        "LELAKI (%)",
        "PEREMPUAN",
        "PEREMPUAN (%)",

        "MELAYU",
        "MELAYU (%)",
        "CINA",
        "CINA (%)",
        "INDIA",
        "INDIA (%)",
        "LAIN-LAIN",
        "LAIN-LAIN (%)",
    ]

    for age_group in age_groups:
        headers.extend([
            age_group,
            f"{age_group} (%)",
        ])

    for party in PARTY_COLS:
        headers.extend([
            party,
            f"{party} (%)",
        ])

    for sikap in SIKAP_COLS:
        headers.extend([
            sikap,
            f"{sikap} (%)",
        ])

    headers.extend([
        "PENGUNDI AWAL",
        "PENGUNDI AWAL (%)",

        "POLIS",
        "POLIS (%)",

        "PASANGAN POLIS",
        "PASANGAN POLIS (%)",

        "ASKAR",
        "ASKAR (%)",

        "PASANGAN ASKAR",
        "PASANGAN ASKAR (%)",
    ])

    return headers


# ============================================================
# BUILD DM ROW
# ============================================================

def build_dm_row(
    kod_dm,
    nama_dm,
    grp,
    age_groups,
):
    total = len(grp)

    row = {
        "KOD DM": format_kod_dm(kod_dm),
        "NAMA DM": nama_dm,
        "JUMLAH": total,
    }

    sex_vc = grp["_jantina"].value_counts()
    race_vc = grp["_race"].value_counts()
    age_vc = grp["_age_group"].value_counts()
    party_vc = grp["_party"].value_counts()
    sikap_vc = grp["_sikap"].value_counts()
    awal_vc = grp["_awal_type"].value_counts()

    # --------------------------------------------------------
    # SEX
    # --------------------------------------------------------

    for key, label in [
        ("L", "LELAKI"),
        ("P", "PEREMPUAN"),
    ]:
        count = sex_vc.get(
            key,
            0,
        )

        row[label] = count
        row[f"{label} (%)"] = pct(
            count,
            total,
        )

    # --------------------------------------------------------
    # RACE
    # --------------------------------------------------------

    for race in MAIN_RACES:
        count = race_vc.get(
            race,
            0,
        )

        row[race] = count
        row[f"{race} (%)"] = pct(
            count,
            total,
        )

    # --------------------------------------------------------
    # AGE
    # --------------------------------------------------------

    for age_group in age_groups:
        count = age_vc.get(
            age_group,
            0,
        )

        row[age_group] = count
        row[f"{age_group} (%)"] = pct(
            count,
            total,
        )

    # --------------------------------------------------------
    # PARTY
    # --------------------------------------------------------

    for party in PARTY_COLS:
        count = party_vc.get(
            party,
            0,
        )

        row[party] = count
        row[f"{party} (%)"] = pct(
            count,
            total,
        )

    # --------------------------------------------------------
    # SIKAP
    # --------------------------------------------------------

    for sikap in SIKAP_COLS:
        count = sikap_vc.get(
            sikap,
            0,
        )

        row[sikap] = count
        row[f"{sikap} (%)"] = pct(
            count,
            total,
        )

    # --------------------------------------------------------
    # PENGUNDI AWAL
    # --------------------------------------------------------

    pengundi_awal = (
        grp["_NoPerkhidmatan_clean"]
        .ne("")
        .sum()
    )

    polis = awal_vc.get(
        "POLIS",
        0,
    )

    askar = awal_vc.get(
        "ASKAR",
        0,
    )

    pasangan_polis = (
        grp["_Pasangan Polis"]
        .sum()
    )

    pasangan_askar = (
        grp["_Pasangan Askar"]
        .sum()
    )

    row["PENGUNDI AWAL"] = pengundi_awal
    row["PENGUNDI AWAL (%)"] = pct(
        pengundi_awal,
        total,
    )

    row["POLIS"] = polis
    row["POLIS (%)"] = pct(
        polis,
        total,
    )

    row["PASANGAN POLIS"] = pasangan_polis
    row["PASANGAN POLIS (%)"] = pct(
        pasangan_polis,
        total,
    )

    row["ASKAR"] = askar
    row["ASKAR (%)"] = pct(
        askar,
        total,
    )

    row["PASANGAN ASKAR"] = pasangan_askar
    row["PASANGAN ASKAR (%)"] = pct(
        pasangan_askar,
        total,
    )

    return row


# ============================================================
# BUILD RUMUSAN
# ============================================================

def build_rumusan_df(
    dun_df,
    age_groups,
    headers,
):
    rows = []

    grouped = dun_df.groupby(
        [
            "_KOD DM",
            "_NAMA DM",
        ],
        dropna=False,
    )

    for (
        kod_dm,
        nama_dm,
    ), grp in grouped:

        rows.append(
            build_dm_row(
                kod_dm,
                nama_dm,
                grp,
                age_groups,
            )
        )

    rumusan_df = pd.DataFrame(
        rows
    )

    for header in headers:

        if header not in rumusan_df.columns:
            rumusan_df[header] = 0

    rumusan_df = rumusan_df[
        headers
    ]

    rumusan_df = rumusan_df.sort_values(
        by="KOD DM",
        kind="stable",
    )

    return add_total_row(
        rumusan_df,
        headers,
    )


# ============================================================
# TOTAL ROW
# ============================================================

def add_total_row(
    df,
    headers,
):
    total = (
        pd.to_numeric(
            df["JUMLAH"],
            errors="coerce",
        )
        .fillna(0)
        .sum()
    )

    total_row = {
        "KOD DM": "",
        "NAMA DM": "",
        "JUMLAH": total,
    }

    for header in headers:

        if header in {
            "KOD DM",
            "NAMA DM",
            "JUMLAH",
        }:
            continue

        if header.endswith("(%)"):

            base = header.replace(
                " (%)",
                "",
            )

            total_row[header] = pct(
                total_row.get(
                    base,
                    0,
                ),
                total,
            )

        else:

            total_row[header] = (
                pd.to_numeric(
                    df[header],
                    errors="coerce",
                )
                .fillna(0)
                .sum()
            )

    return pd.concat(
        [
            df,
            pd.DataFrame([
                total_row
            ]),
        ],
        ignore_index=True,
    )


# ============================================================
# EXCEL FORMATTING
# ============================================================

def write_dun_sheet(
    ws,
    rumusan_df,
    headers,
    age_groups,
):
    BLUE = "9DC3E6"
    GREEN = "A9D18E"
    ORANGE = "F4B183"
    YELLOW = "FFD966"
    PURPLE = "B4A7D6"
    GREY = "D9D9D9"

    thin = Side(
        style="thin",
        color="000000",
    )

    medium = Side(
        style="medium",
        color="000000",
    )

    # --------------------------------------------------------
    # DETERMINE COLUMN GROUPS
    # --------------------------------------------------------

    header_positions = {
        header: index
        for index, header
        in enumerate(headers, start=1)
    }

    sex_start = header_positions["LELAKI"]
    sex_end = header_positions["PEREMPUAN (%)"]

    race_start = header_positions["MELAYU"]
    race_end = header_positions["LAIN-LAIN (%)"]

    age_start = header_positions[
        age_groups[0]
    ]

    age_end = header_positions[
        f"{age_groups[-1]} (%)"
    ]

    party_start = header_positions[
        PARTY_COLS[0]
    ]

    party_end = header_positions[
        f"{PARTY_COLS[-1]} (%)"
    ]

    sikap_start = header_positions[
        SIKAP_COLS[0]
    ]

    sikap_end = header_positions[
        f"{SIKAP_COLS[-1]} (%)"
    ]

    special_start = header_positions[
        "PENGUNDI AWAL"
    ]

    special_end = header_positions[
        "PASANGAN ASKAR (%)"
    ]

    # --------------------------------------------------------
    # FILL COLOURS
    # --------------------------------------------------------

    group_fill = {}

    for column in range(
        sex_start,
        sex_end + 1,
    ):
        group_fill[column] = BLUE

    for column in range(
        race_start,
        race_end + 1,
    ):
        group_fill[column] = GREEN

    for column in range(
        age_start,
        age_end + 1,
    ):
        group_fill[column] = ORANGE

    for column in range(
        party_start,
        party_end + 1,
    ):
        group_fill[column] = YELLOW

    for column in range(
        sikap_start,
        sikap_end + 1,
    ):
        group_fill[column] = GREY

    for column in range(
        special_start,
        special_end + 1,
    ):
        group_fill[column] = PURPLE

    # --------------------------------------------------------
    # GROUP EDGES
    # --------------------------------------------------------

    group_left_edges = {
        1,
        sex_start,
        race_start,
        age_start,
        party_start,
        sikap_start,
        special_start,
    }

    group_right_edges = {
        2,
        sex_end,
        race_end,
        age_end,
        party_end,
        sikap_end,
        special_end,
    }

    # --------------------------------------------------------
    # HEADER
    # --------------------------------------------------------

    for column, header in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(
            row=1,
            column=column,
            value=header,
        )

        cell.font = Font(
            name="Calibri",
            size=11,
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        if column in group_fill:
            cell.fill = PatternFill(
                "solid",
                fgColor=group_fill[column],
            )

        left = (
            medium
            if column in group_left_edges
            else thin
        )

        right = (
            medium
            if column in group_right_edges
            else thin
        )

        cell.border = Border(
            left=left,
            right=right,
            top=medium,
            bottom=medium,
        )

    # --------------------------------------------------------
    # DATA
    # --------------------------------------------------------

    total_row_number = (
        len(rumusan_df) + 1
    )

    for row_number, row in enumerate(
        rumusan_df.itertuples(
            index=False
        ),
        start=2,
    ):
        is_total = (
            row_number
            == total_row_number
        )

        for column, value in enumerate(
            row,
            start=1,
        ):
            cell = ws.cell(
                row=row_number,
                column=column,
                value=value,
            )

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column == 2
                    else "center"
                ),
                vertical="center",
            )

            cell.font = Font(
                name="Calibri",
                size=11,
                bold=is_total,
            )

            if (
                is_total
                and column in group_fill
            ):
                cell.fill = PatternFill(
                    "solid",
                    fgColor=group_fill[column],
                )

            left = (
                medium
                if column in group_left_edges
                else thin
            )

            right = (
                medium
                if column in group_right_edges
                else thin
            )

            cell.border = Border(
                left=left,
                right=right,
                top=(
                    medium
                    if is_total
                    else thin
                ),
                bottom=(
                    medium
                    if is_total
                    else thin
                ),
            )

            header = headers[
                column - 1
            ]

            if isinstance(
                value,
                (int, float),
            ):
                if "(%)" in header:
                    cell.number_format = "0.0"
                else:
                    cell.number_format = "#,##0"

    # --------------------------------------------------------
    # COLUMN WIDTHS
    # --------------------------------------------------------

    widths = {
        "A": 13,
        "B": 25,
        "C": 13,
        "D": 11.3,
        "E": 14.7,
        "F": 17,
        "G": 20.6,
        "H": 13.1,
        "I": 16.6,
        "J": 10,
        "K": 13.4,
        "L": 10.7,
        "M": 14.1,
        "N": 14.6,
        "O": 18.1,
        "P": 10.3,
        "Q": 13.7,
        "R": 10.3,
        "S": 13.7,
        "T": 10.3,
        "U": 13.7,
        "V": 10.3,
        "W": 13.7,
        "X": 10.3,
        "Y": 13.7,
        "Z": 8.6,
        "AA": 12,
        "AB": 9,
        "AC": 12.4,
        "AD": 9,
        "AE": 12.4,
        "AF": 10.9,
        "AG": 14.3,
        "AH": 11.7,
        "AI": 15.1,
        "AJ": 11,
        "AK": 14.4,
        "AL": 12.4,
        "AM": 15.9,
        "AN": 11.6,
        "AO": 15,
        "AP": 21.3,
        "AQ": 24.9,
        "AR": 10.6,
        "AS": 14,
        "AT": 21.4,
        "AU": 25,
        "AV": 11.4,
        "AW": 14.9,
        "AX": 22.4,
        "AY": 26,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    # NAMA DM width
    max_len = 0

    for row_number in range(
        1,
        ws.max_row + 1,
    ):
        value = ws.cell(
            row=row_number,
            column=2,
        ).value

        max_len = max(
            max_len,
            len(str(value or "")),
        )

    ws.column_dimensions[
        "B"
    ].width = min(
        max(max_len + 4, 25),
        60,
    )

    # --------------------------------------------------------
    # ROW / FREEZE
    # --------------------------------------------------------

    ws.row_dimensions[1].height = 15.75

    ws.freeze_panes = "A2"

    # --------------------------------------------------------
    # EXCEL TABLE
    # --------------------------------------------------------

    end_row = (
        len(rumusan_df) + 1
    )

    end_column = len(headers)

    end_column_letter = (
        ws.cell(
            row=1,
            column=end_column,
        ).column_letter
    )

    table_ref = (
        f"A1:{end_column_letter}{end_row}"
    )

    safe_title = re.sub(
        r"[^A-Za-z0-9_]",
        "_",
        ws.title,
    )

    table_name = (
        f"Table_{safe_title}"
    )

    if not re.match(
        r"^[A-Za-z_]",
        table_name,
    ):
        table_name = (
            f"Table_{table_name}"
        )

    table_name = table_name[
        :250
    ]

    table = Table(
        displayName=table_name,
        ref=table_ref,
    )

    table_style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )

    table.tableStyleInfo = table_style

    ws.add_table(table)


# ============================================================
# MAIN GENERATOR
# ============================================================

def generate_demografik(
    uploaded_files,
    age_groups=None,
):
    logs = []
    all_data = []

    # --------------------------------------------------------
    # AGE GROUPS
    # --------------------------------------------------------

    age_groups, parsed_age_groups = parse_age_groups(
        age_groups
    )

    headers = build_headers(
        age_groups
    )

    # --------------------------------------------------------
    # LOAD FILES
    # --------------------------------------------------------

    for uploaded_file in uploaded_files:

        filename = uploaded_file.name

        try:
            uploaded_file.seek(0)

            df = pd.read_excel(
                uploaded_file,
                dtype=str,
            )

            df.columns = [
                str(column).strip()
                for column in df.columns
            ]

            # ------------------------------------------------
            # COLUMN DETECTION
            # ------------------------------------------------

            col_dm = get_col(
                df,
                [
                    "KOD DM",
                    "kod_dm",
                ],
            )

            col_nama_dm = get_col(
                df,
                [
                    "NamaDM",
                    "nama_dm",
                    "NAMA DM",
                ],
            )

            col_kod_dun = get_col(
                df,
                [
                    "kod_dun",
                    "KOD DUN",
                    "KODDUN",
                ],
            )

            col_nama_dun = get_col(
                df,
                [
                    "nama_dun",
                    "DUN",
                    "NAMA DUN",
                ],
            )

            col_jantina = get_col(
                df,
                [
                    "JANTINA",
                    "jantina",
                ],
            )

            col_bangsa = get_col(
                df,
                [
                    "kaum",
                    "BANGSA",
                    "kategori_kaum",
                ],
            )

            col_umur = get_col(
                df,
                [
                    "UMUR",
                    "umur",
                ],
            )

            col_party = get_col(
                df,
                [
                    "party",
                    "PARTY",
                ],
            )

            col_sikap = get_col(
                df,
                [
                    "CATATAN",
                    "sikap",
                ],
            )

            col_no = get_col(
                df,
                [
                    "NoPerkhidmatan",
                    "noperkhidmatan",
                ],
            )

            col_pasangan = get_col(
                df,
                [
                    "NoKPPasangan",
                    "NoPerkhidmatanPasangan",
                    "noperkhidmatanpasangan",
                ],
            )

            required = {
                "KOD DM": col_dm,
                "NamaDM": col_nama_dm,
                "kod_dun": col_kod_dun,
                "nama_dun / DUN": col_nama_dun,
                "JANTINA": col_jantina,
                "BANGSA": col_bangsa,
                "UMUR": col_umur,
                "NoPerkhidmatan": col_no,
                "NoKPPasangan": col_pasangan,
            }

            missing = [
                name
                for name, value
                in required.items()
                if value is None
            ]

            if missing:
                logs.append(
                    f"Skipped {filename} — "
                    f"missing columns: {missing}"
                )
                continue

            # ------------------------------------------------
            # NORMALISE DATA
            # ------------------------------------------------

            df["_KOD_DUN"] = (
                df[col_kod_dun]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            df["_NAMA_DUN"] = (
                df[col_nama_dun]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
            )

            df["_KOD DM"] = (
                df[col_dm]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            df["_NAMA DM"] = (
                df[col_nama_dm]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            df["_jantina"] = (
                df[col_jantina]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
            )

            df["_race"] = (
                df[col_bangsa]
                .apply(normalise_race)
            )

            df["_age_group"] = (
                df[col_umur]
                .apply(
                    lambda value:
                    get_age_group(
                        value,
                        parsed_age_groups,
                    )
                )
            )

            if col_party:
                df["_party"] = (
                    df[col_party]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .str.upper()
                )
            else:
                df["_party"] = ""

            if col_sikap:
                df["_sikap"] = (
                    df[col_sikap]
                    .apply(normalise_sikap)
                )
            else:
                df["_sikap"] = ""

            df["_NoPerkhidmatan_clean"] = (
                df[col_no]
                .apply(clean_service_no)
            )

            df["_NoKPPasangan_clean"] = (
                df[col_pasangan]
                .apply(clean_service_no)
            )

            df["_awal_type"] = (
                df["_NoPerkhidmatan_clean"]
                .apply(classify_awal)
            )

            df["_Pasangan Polis"] = (
                df["_NoKPPasangan_clean"]
                .apply(
                    lambda value:
                    1 if is_polis(value)
                    else 0
                )
            )

            df["_Pasangan Askar"] = (
                df["_NoKPPasangan_clean"]
                .apply(
                    lambda value:
                    1 if is_askar(value)
                    else 0
                )
            )

            all_data.append(df)

            logs.append(
                f"Loaded {filename}: "
                f"{len(df):,} rows"
            )

        except Exception as error:

            logs.append(
                f"Error reading "
                f"{filename}: {error}"
            )

    # --------------------------------------------------------
    # VALIDATE
    # --------------------------------------------------------

    if not all_data:
        raise ValueError(
            "No valid data loaded.\n"
            + "\n".join(logs)
        )

    final_df = pd.concat(
        all_data,
        ignore_index=True,
    )

    # --------------------------------------------------------
    # DUN LIST
    # --------------------------------------------------------

    dun_combos = (
        final_df[
            [
                "_KOD_DUN",
                "_NAMA_DUN",
            ]
        ]
        .drop_duplicates()
    )

    dun_combos = dun_combos[
        (
            dun_combos["_KOD_DUN"]
            != ""
        )
        &
        (
            dun_combos["_NAMA_DUN"]
            != ""
        )
    ]

    dun_combos = sorted(
        dun_combos.itertuples(
            index=False,
            name=None,
        ),
        key=lambda pair:
        kod_dun_sort_key(pair[0]),
    )

    if not dun_combos:
        raise ValueError(
            "No DUN name/kod_dun found "
            "in the uploaded data."
        )

    # --------------------------------------------------------
    # WORKBOOK
    # --------------------------------------------------------

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    existing_sheet_names = set()

    # --------------------------------------------------------
    # ONE SHEET PER DUN
    # --------------------------------------------------------

    for (
        kod_dun,
        nama_dun,
    ) in dun_combos:

        dun_df = final_df[
            (
                final_df["_KOD_DUN"]
                == kod_dun
            )
            &
            (
                final_df["_NAMA_DUN"]
                == nama_dun
            )
        ]

        rumusan_df = build_rumusan_df(
            dun_df,
            age_groups,
            headers,
        )

        sheet_label = format_sheet_label(
            kod_dun,
            nama_dun,
        )

        sheet_name = clean_sheet_name(
            sheet_label,
            existing_sheet_names,
        )

        worksheet = workbook.create_sheet(
            title=sheet_name
        )

        write_dun_sheet(
            worksheet,
            rumusan_df,
            headers,
            age_groups,
        )

        logs.append(
            f"Built sheet "
            f"'{sheet_name}': "
            f"{len(dun_df):,} rows, "
            f"{len(rumusan_df) - 1} DM(s)"
        )

    # --------------------------------------------------------
    # SAVE
    # --------------------------------------------------------

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    # --------------------------------------------------------
    # OUTPUT NAME
    # --------------------------------------------------------

    if len(dun_combos) == 1:

        kod_dun, nama_dun = (
            dun_combos[0]
        )

        out_name = (
            "DEMOGRAFIK "
            + clean_filename(
                format_sheet_label(
                    kod_dun,
                    nama_dun,
                )
            )
            + ".xlsx"
        )

    else:

        out_name = (
            "DEMOGRAFIK "
            f"({len(dun_combos)} DUN).xlsx"
        )

    return (
        output.getvalue(),
        out_name,
        logs,
    )
