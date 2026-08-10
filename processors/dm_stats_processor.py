import io
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Border,
    Side,
    PatternFill,
    Alignment
)


# ============================================================
# CONSTANTS
# ============================================================

MAIN_RACES = [
    "MELAYU",
    "CINA",
    "INDIA",
    "LAIN-LAIN"
]

PARTY_COLS = [
    "UMNO",
    "PKR",
    "PAS",
    "PPBM"
]

SIKAP_COLS = [
    "PUTIH",
    "KELABU",
    "HITAM"
]


# ============================================================
# COLUMN HELPERS
# ============================================================

def get_col(df, possible_names):

    col_map = {
        str(c).lower().strip(): c
        for c in df.columns
    }

    for name in possible_names:

        key = str(name).lower().strip()

        if key in col_map:
            return col_map[key]

    return None


# ============================================================
# CLEANING
# ============================================================

def clean_service_no(value):

    if pd.isna(value):
        return ""

    n = str(value).strip().upper()

    if n in {
        "",
        "NAN",
        "NONE",
        "NULL"
    }:
        return ""

    return n


def clean_filename(value):

    name = str(value).strip().upper()

    name = re.sub(
        r'[\\/:*?"<>|]',
        " ",
        name
    )

    name = " ".join(name.split())

    return name if name else "OUTPUT"


# ============================================================
# KOD HELPERS
# ============================================================

def digits_only(value):

    if pd.isna(value):
        return ""

    value = str(value).strip()

    if value.lower() in {
        "",
        "nan",
        "none",
        "null"
    }:
        return ""

    return re.sub(
        r"\D",
        "",
        value.split(".")[0]
    )


def numeric_sort_key(value):

    digits = digits_only(value)

    if not digits:
        return -1

    try:
        return int(digits)
    except Exception:
        return -1


def format_kod_dm(value):

    kod = digits_only(value)

    if not kod:
        return ""

    kod = kod.zfill(7)

    return (
        f"{kod[:3]}/"
        f"{kod[3:5]}/"
        f"{kod[5:]}"
    )


def format_kod_dun(value):

    kod = digits_only(value)

    if not kod:
        return ""

    kod = kod.zfill(5)

    return (
        f"{kod[:3]}/"
        f"{kod[3:]}"
    )


def format_kod_parlimen(value):

    kod = digits_only(value)

    if not kod:
        return ""

    return kod.zfill(3)


def get_dun_code_from_dm(kod_dm):

    kod = digits_only(kod_dm)

    if not kod:
        return ""

    kod = kod.zfill(7)

    return (
        f"{kod[:3]}/"
        f"{kod[3:5]}"
    )


# ============================================================
# NORMALISATION
# ============================================================

def normalise_race(value):

    if pd.isna(value):
        return "LAIN-LAIN"

    race = str(value).strip().upper()

    if race in {
        "MELAYU",
        "CINA",
        "INDIA"
    }:
        return race

    return "LAIN-LAIN"


def normalise_sikap(value):

    if pd.isna(value):
        return ""

    sikap = str(value).strip().upper()

    if sikap in {
        "KELABU-LAMA",
        "KELABU-BARU"
    }:
        return "KELABU"

    if sikap in {
        "PUTIH",
        "KELABU",
        "HITAM"
    }:
        return sikap

    return ""


# ============================================================
# SERVICE CLASSIFICATION
# ============================================================

def classify_awal(value):

    n = clean_service_no(value)

    if not n:
        return ""

    if n.startswith("G") or n.startswith("RF"):
        return "POLIS"

    if n.startswith("T"):
        return "ASKAR"

    return "PENGUNDI AWAL"


def is_polis(value):

    n = clean_service_no(value)

    return (
        n.startswith("G")
        or n.startswith("RF")
    )


def is_askar(value):

    n = clean_service_no(value)

    return n.startswith("T")


# ============================================================
# AGE
# ============================================================

def parse_age(value):

    try:

        if pd.isna(value):
            return None

        age = int(float(value))

        return age

    except Exception:

        return None


def get_age_group(value, age_groups):

    age = parse_age(value)

    if age is None:
        return ""

    for label in age_groups:

        label_clean = str(label).strip()

        # ----------------------------------------------------
        # Example: 61+
        # ----------------------------------------------------

        plus_match = re.fullmatch(
            r"(\d+)\s*\+",
            label_clean
        )

        if plus_match:

            minimum = int(
                plus_match.group(1)
            )

            if age >= minimum:
                return label_clean

            continue

        # ----------------------------------------------------
        # Example: 18-21
        # ----------------------------------------------------

        range_match = re.fullmatch(
            r"(\d+)\s*-\s*(\d+)",
            label_clean
        )

        if range_match:

            minimum = int(
                range_match.group(1)
            )

            maximum = int(
                range_match.group(2)
            )

            if minimum <= age <= maximum:
                return label_clean

    return ""


# ============================================================
# PERCENTAGE
# ============================================================

def pct(part, total):

    if not total:
        return 0

    return round(
        (part / total) * 100,
        1
    )


# ============================================================
# HEADERS
# ============================================================

def build_headers(age_groups):

    headers = [
        "KOD",
        "NAMA",
        "JUMLAH",

        "LELAKI",
        "LELAKI (%)",

        "PEREMPUAN",
        "PEREMPUAN (%)",

        "MELAYU",
        "MELAYU (%)",

        "CINA",
        "CINA (%)",

        "INDIA",
        "INDIA (%)",

        "LAIN-LAIN",
        "LAIN-LAIN (%)"
    ]

    for age in age_groups:

        headers.append(age)
        headers.append(f"{age} (%)")

    headers.extend([
        "UMNO",
        "UMNO (%)",

        "PKR",
        "PKR (%)",

        "PAS",
        "PAS (%)",

        "PPBM",
        "PPBM (%)",

        "PUTIH",
        "PUTIH (%)",

        "KELABU",
        "KELABU (%)",

        "HITAM",
        "HITAM (%)",

        "PENGUNDI AWAL",
        "PENGUNDI AWAL (%)",

        "POLIS",
        "POLIS (%)",

        "PASANGAN POLIS",
        "PASANGAN POLIS (%)",

        "ASKAR",
        "ASKAR (%)",

        "PASANGAN ASKAR",
        "PASANGAN ASKAR (%)"
    ])

    return headers


# ============================================================
# BUILD ONE SUMMARY ROW
# ============================================================

def build_summary_row(
    kod,
    nama,
    grp,
    age_groups
):

    total = len(grp)

    row = {
        "KOD": kod,
        "NAMA": nama,
        "JUMLAH": total
    }

    # ========================================================
    # SEX
    # ========================================================

    sex_vc = grp["_jantina"].value_counts()

    for key, label in [
        ("L", "LELAKI"),
        ("P", "PEREMPUAN")
    ]:

        count = sex_vc.get(
            key,
            0
        )

        row[label] = count
        row[f"{label} (%)"] = pct(
            count,
            total
        )

    # ========================================================
    # RACE
    # ========================================================

    race_vc = grp["_race"].value_counts()

    for race in MAIN_RACES:

        count = race_vc.get(
            race,
            0
        )

        row[race] = count
        row[f"{race} (%)"] = pct(
            count,
            total
        )

    # ========================================================
    # AGE
    # ========================================================

    age_vc = grp["_age_group"].value_counts()

    for age in age_groups:

        count = age_vc.get(
            age,
            0
        )

        row[age] = count
        row[f"{age} (%)"] = pct(
            count,
            total
        )

    # ========================================================
    # PARTY
    # ========================================================

    party_vc = grp["_party"].value_counts()

    for party in PARTY_COLS:

        count = party_vc.get(
            party,
            0
        )

        row[party] = count
        row[f"{party} (%)"] = pct(
            count,
            total
        )

    # ========================================================
    # SIKAP
    # ========================================================

    sikap_vc = grp["_sikap"].value_counts()

    for sikap in SIKAP_COLS:

        count = sikap_vc.get(
            sikap,
            0
        )

        row[sikap] = count
        row[f"{sikap} (%)"] = pct(
            count,
            total
        )

    # ========================================================
    # PENGUNDI AWAL / POLIS / ASKAR
    # ========================================================

    awal_vc = grp["_awal_type"].value_counts()

    pengundi_awal = (
        grp["_NoPerkhidmatan_clean"]
        .ne("")
        .sum()
    )

    polis = awal_vc.get(
        "POLIS",
        0
    )

    askar = awal_vc.get(
        "ASKAR",
        0
    )

    pasangan_polis = int(
        grp["_Pasangan Polis"].sum()
    )

    pasangan_askar = int(
        grp["_Pasangan Askar"].sum()
    )

    row["PENGUNDI AWAL"] = pengundi_awal
    row["PENGUNDI AWAL (%)"] = pct(
        pengundi_awal,
        total
    )

    row["POLIS"] = polis
    row["POLIS (%)"] = pct(
        polis,
        total
    )

    row["PASANGAN POLIS"] = pasangan_polis
    row["PASANGAN POLIS (%)"] = pct(
        pasangan_polis,
        total
    )

    row["ASKAR"] = askar
    row["ASKAR (%)"] = pct(
        askar,
        total
    )

    row["PASANGAN ASKAR"] = pasangan_askar
    row["PASANGAN ASKAR (%)"] = pct(
        pasangan_askar,
        total
    )

    return row


# ============================================================
# BUILD SUMMARY TABLE
# ============================================================

def build_summary_df(
    df,
    group_columns,
    age_groups,
    headers,
    kod_formatter=None
):

    rows = []

    grouped = df.groupby(
        group_columns,
        dropna=False,
        sort=False
    )

    for group_values, grp in grouped:

        if not isinstance(
            group_values,
            tuple
        ):
            group_values = (
                group_values,
            )

        kod = group_values[0]
        nama = group_values[1]

        if pd.isna(kod):
            kod = ""

        if pd.isna(nama):
            nama = ""

        kod = str(kod).strip()
        nama = str(nama).strip()

        if kod_formatter:
            kod = kod_formatter(kod)

        rows.append(
            build_summary_row(
                kod,
                nama,
                grp,
                age_groups
            )
        )

    summary_df = pd.DataFrame(rows)

    # ========================================================
    # Ensure all headers exist
    # ========================================================

    for header in headers:

        if header not in summary_df.columns:
            summary_df[header] = 0

    summary_df = summary_df[headers]

    # ========================================================
    # Sort by KOD
    # ========================================================

    if not summary_df.empty:

        summary_df["_sort"] = (
            summary_df["KOD"]
            .apply(numeric_sort_key)
        )

        summary_df = (
            summary_df
            .sort_values(
                "_sort",
                kind="stable"
            )
            .drop(columns="_sort")
            .reset_index(drop=True)
        )

    return summary_df


# ============================================================
# TOTAL ROW FROM SUMMARY DATAFRAME
#
# Used only for PARLIMEN.
# DUN totals inside DM are calculated from underlying records.
# ============================================================

def add_total_row(
    df,
    headers
):

    if df.empty:

        total = 0

    else:

        total = pd.to_numeric(
            df["JUMLAH"],
            errors="coerce"
        ).fillna(0).sum()

    total_row = {
        "KOD": "",
        "NAMA": "",
        "JUMLAH": total
    }

    for header in headers:

        if header in {
            "KOD",
            "NAMA",
            "JUMLAH"
        }:
            continue

        if header.endswith("(%)"):

            base = header.replace(
                " (%)",
                ""
            )

            numerator = total_row.get(
                base,
                0
            )

            total_row[header] = pct(
                numerator,
                total
            )

        else:

            if header in df.columns:

                total_row[header] = (
                    pd.to_numeric(
                        df[header],
                        errors="coerce"
                    )
                    .fillna(0)
                    .sum()
                )

            else:

                total_row[header] = 0

    return pd.concat(
        [
            df,
            pd.DataFrame([total_row])
        ],
        ignore_index=True
    )


# ============================================================
# CALCULATE TOTAL FROM UNDERLYING DUN RECORDS
#
# IMPORTANT:
# Percentages are recalculated from the underlying records.
# They are NOT averaged from DM percentages.
# ============================================================

def build_dun_total_row(
    dun_code,
    dun_name,
    dun_records,
    age_groups
):

    return build_summary_row(
        dun_code,
        dun_name,
        dun_records,
        age_groups
    )


# ============================================================
# SHEET FORMATTING HELPERS
# ============================================================

def get_group_fills(headers):

    BLUE = "9DC3E6"
    GREEN = "A9D18E"
    ORANGE = "F4B183"
    YELLOW = "FFD966"
    WHITE = "D9D9D9"
    PURPLE = "B4A7D6"

    group_fill = {}

    # --------------------------------------------------------
    # SEX
    # --------------------------------------------------------

    for c in range(4, 8):
        group_fill[c] = BLUE

    # --------------------------------------------------------
    # RACE
    # --------------------------------------------------------

    for c in range(8, 16):
        group_fill[c] = GREEN

    # --------------------------------------------------------
    # AGE
    # --------------------------------------------------------

    age_start = 16

    for index, header in enumerate(
        headers,
        start=1
    ):

        if index < age_start:
            continue

        if header in {
            "UMNO",
            "UMNO (%)"
        }:
            break

        group_fill[index] = ORANGE

    # --------------------------------------------------------
    # PARTY
    # --------------------------------------------------------

    try:

        party_start = (
            headers.index("UMNO") + 1
        )

        for c in range(
            party_start,
            party_start + 8
        ):
            group_fill[c] = YELLOW

    except ValueError:

        party_start = None

    # --------------------------------------------------------
    # SIKAP
    # --------------------------------------------------------

    try:

        sikap_start = (
            headers.index("PUTIH") + 1
        )

        for c in range(
            sikap_start,
            sikap_start + 6
        ):
            group_fill[c] = WHITE

    except ValueError:

        sikap_start = None

    # --------------------------------------------------------
    # AWAL
    # --------------------------------------------------------

    try:

        awal_start = (
            headers.index(
                "PENGUNDI AWAL"
            ) + 1
        )

        for c in range(
            awal_start,
            len(headers) + 1
        ):
            group_fill[c] = PURPLE

    except ValueError:

        awal_start = None

    return group_fill


def get_group_edges(headers):

    group_left_edges = {
        1,
        4,
        8,
        16
    }

    group_right_edges = {
        2,
        7,
        15
    }

    try:

        party_start = (
            headers.index("UMNO") + 1
        )

        group_left_edges.add(
            party_start
        )

        group_right_edges.add(
            party_start + 7
        )

    except ValueError:

        pass

    try:

        sikap_start = (
            headers.index("PUTIH") + 1
        )

        group_left_edges.add(
            sikap_start
        )

        group_right_edges.add(
            sikap_start + 5
        )

    except ValueError:

        pass

    try:

        awal_start = (
            headers.index(
                "PENGUNDI AWAL"
            ) + 1
        )

        group_left_edges.add(
            awal_start
        )

        group_right_edges.add(
            len(headers)
        )

    except ValueError:

        pass

    return (
        group_left_edges,
        group_right_edges
    )


# ============================================================
# WRITE MAIN HEADER
# ============================================================

def write_main_header(
    ws,
    headers,
    row_number=1
):

    group_fill = get_group_fills(
        headers
    )

    (
        group_left_edges,
        group_right_edges
    ) = get_group_edges(
        headers
    )

    thin = Side(
        style="thin",
        color="000000"
    )

    medium = Side(
        style="medium",
        color="000000"
    )

    for col_idx, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=row_number,
            column=col_idx,
            value=header
        )

        cell.font = Font(
            name="Calibri",
            size=11,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        if col_idx in group_fill:

            cell.fill = PatternFill(
                "solid",
                fgColor=group_fill[col_idx]
            )

        left = (
            medium
            if col_idx in group_left_edges
            else thin
        )

        right = (
            medium
            if col_idx in group_right_edges
            else thin
        )

        cell.border = Border(
            left=left,
            right=right,
            top=medium,
            bottom=medium
        )

    ws.row_dimensions[
        row_number
    ].height = 30


# ============================================================
# WRITE DATA ROW
#
# NO FILL COLOR ON NORMAL DM ROWS.
# ============================================================

def write_data_row(
    ws,
    row_number,
    values,
    headers,
    bold=False
):

    thin = Side(
        style="thin",
        color="000000"
    )

    medium = Side(
        style="medium",
        color="000000"
    )

    (
        group_left_edges,
        group_right_edges
    ) = get_group_edges(
        headers
    )

    for col_idx, value in enumerate(
        values,
        start=1
    ):

        cell = ws.cell(
            row=row_number,
            column=col_idx,
            value=value
        )

        # ----------------------------------------------------
        # Alignment
        # ----------------------------------------------------

        if col_idx == 2:

            cell.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

        else:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        cell.font = Font(
            name="Calibri",
            size=11,
            bold=bold
        )

        # ----------------------------------------------------
        # IMPORTANT:
        # NO FILL FOR NORMAL DATA ROWS
        # ----------------------------------------------------

        cell.fill = PatternFill(
            fill_type=None
        )

        # ----------------------------------------------------
        # Borders
        # ----------------------------------------------------

        left = (
            medium
            if col_idx in group_left_edges
            else thin
        )

        right = (
            medium
            if col_idx in group_right_edges
            else thin
        )

        cell.border = Border(
            left=left,
            right=right,
            top=thin,
            bottom=thin
        )

        # ----------------------------------------------------
        # NUMBER FORMAT
        # ----------------------------------------------------

        header = headers[
            col_idx - 1
        ]

        if header.endswith("(%)"):

            cell.number_format = "0.0"

        elif isinstance(
            value,
            (int, float)
        ):

            cell.number_format = "#,##0"

    ws.row_dimensions[
        row_number
    ].height = 22


# ============================================================
# WRITE DUN GROUP HEADER
# ============================================================

def write_dun_group_header(
    ws,
    row_number,
    dun_code,
    dun_name,
    headers
):

    medium = Side(
        style="medium",
        color="000000"
    )

    # --------------------------------------------------------
    # Header fill
    # --------------------------------------------------------

    fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    # --------------------------------------------------------
    # KOD
    # --------------------------------------------------------

    kod_cell = ws.cell(
        row=row_number,
        column=1,
        value=dun_code
    )

    kod_cell.font = Font(
        name="Calibri",
        size=12,
        bold=True
    )

    kod_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # --------------------------------------------------------
    # NAMA
    # --------------------------------------------------------

    nama_cell = ws.cell(
        row=row_number,
        column=2,
        value=dun_name
    )

    nama_cell.font = Font(
        name="Calibri",
        size=12,
        bold=True
    )

    nama_cell.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    # --------------------------------------------------------
    # Remaining cells
    # --------------------------------------------------------

    for col_idx in range(
        1,
        len(headers) + 1
    ):

        cell = ws.cell(
            row=row_number,
            column=col_idx
        )

        cell.fill = fill

        cell.border = Border(
            top=medium,
            bottom=medium,
            left=medium,
            right=medium
        )

    ws.row_dimensions[
        row_number
    ].height = 25


# ============================================================
# WRITE DUN TOTAL ROW
# ============================================================

def write_dun_total_row(
    ws,
    row_number,
    total_row,
    headers
):

    group_fill = get_group_fills(
        headers
    )

    (
        group_left_edges,
        group_right_edges
    ) = get_group_edges(
        headers
    )

    thin = Side(
        style="thin",
        color="000000"
    )

    medium = Side(
        style="medium",
        color="000000"
    )

    for col_idx, header in enumerate(
        headers,
        start=1
    ):

        value = total_row.get(
            header,
            0
        )

        cell = ws.cell(
            row=row_number,
            column=col_idx,
            value=value
        )

        # ----------------------------------------------------
        # Alignment
        # ----------------------------------------------------

        if col_idx == 2:

            cell.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

        else:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        cell.font = Font(
            name="Calibri",
            size=11,
            bold=True
        )

        # ----------------------------------------------------
        # Total row is coloured
        # ----------------------------------------------------

        if col_idx in group_fill:

            cell.fill = PatternFill(
                "solid",
                fgColor=group_fill[col_idx]
            )

        else:

            cell.fill = PatternFill(
                "solid",
                fgColor="E7E6E6"
            )

        # ----------------------------------------------------
        # Borders
        # ----------------------------------------------------

        left = (
            medium
            if col_idx in group_left_edges
            else thin
        )

        right = (
            medium
            if col_idx in group_right_edges
            else thin
        )

        cell.border = Border(
            left=left,
            right=right,
            top=medium,
            bottom=medium
        )

        # ----------------------------------------------------
        # NUMBER FORMAT
        # ----------------------------------------------------

        if header.endswith("(%)"):

            cell.number_format = "0.0"

        elif isinstance(
            value,
            (int, float)
        ):

            cell.number_format = "#,##0"

    ws.row_dimensions[
        row_number
    ].height = 25


# ============================================================
# COLUMN WIDTHS
# ============================================================

def apply_column_widths(ws):

    widths = {
        "A": 13,
        "B": 25,
        "C": 13,

        "D": 11.3,
        "E": 14.7,
        "F": 17,
        "G": 20.6,

        "H": 13.1,
        "I": 16.6,
        "J": 10,
        "K": 13.4,
        "L": 10.7,
        "M": 14.1,
        "N": 14.6,
        "O": 18.1,

        "P": 10.3,
        "Q": 13.7,
        "R": 10.3,
        "S": 13.7,
        "T": 10.3,
        "U": 13.7,
        "V": 10.3,
        "W": 13.7,
        "X": 10.3,
        "Y": 13.7,
        "Z": 8.6,
        "AA": 12,

        "AB": 9,
        "AC": 12.4,
        "AD": 9,
        "AE": 12.4,
        "AF": 10.9,
        "AG": 14.3,
        "AH": 11.7,
        "AI": 15.1,

        "AJ": 11,
        "AK": 14.4,
        "AL": 12.4,
        "AM": 15.9,
        "AN": 11.6,
        "AO": 15,

        "AP": 21.3,
        "AQ": 24.9,
        "AR": 10.6,
        "AS": 14,
        "AT": 21.4,
        "AU": 25,

        "AV": 11.4,
        "AW": 14.9,
        "AX": 22.4,
        "AY": 26
    }

    for column, width in widths.items():

        ws.column_dimensions[
            column
        ].width = width


# ============================================================
# GENERAL WORKSHEET SETTINGS
# ============================================================

def apply_sheet_settings(ws):

    # --------------------------------------------------------
    # Dynamic NAMA width
    # --------------------------------------------------------

    max_len = 0

    for row_idx in range(
        1,
        ws.max_row + 1
    ):

        value = ws.cell(
            row=row_idx,
            column=2
        ).value

        max_len = max(
            max_len,
            len(str(value or ""))
        )

    ws.column_dimensions[
        "B"
    ].width = min(
        max(max_len + 4, 25),
        60
    )

    # --------------------------------------------------------
    # Freeze
    # --------------------------------------------------------

    ws.freeze_panes = "A2"

    # --------------------------------------------------------
    # Page setup
    # --------------------------------------------------------

    ws.page_setup.orientation = "landscape"

    ws.page_setup.paperSize = (
        ws.PAPERSIZE_A4
    )

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2


# ============================================================
# WRITE STANDARD SUMMARY SHEET
#
# Used for PARLIMEN only.
# ============================================================

def write_summary_sheet(
    ws,
    summary_df,
    headers
):

    write_main_header(
        ws,
        headers,
        row_number=1
    )

    for row_idx, row in enumerate(
        summary_df.itertuples(
            index=False,
            name=None
        ),
        start=2
    ):

        write_data_row(
            ws,
            row_idx,
            row,
            headers,
            bold=(
                row_idx == len(summary_df) + 1
            )
        )

    # --------------------------------------------------------
    # Colour final total row
    # --------------------------------------------------------

    if not summary_df.empty:

        total_row_number = (
            len(summary_df) + 1
        )

        total_dict = (
            summary_df
            .iloc[-1]
            .to_dict()
        )

        write_dun_total_row(
            ws,
            total_row_number,
            total_dict,
            headers
        )

    apply_column_widths(ws)
    apply_sheet_settings(ws)


# ============================================================
# WRITE COMBINED DM + DUN WORKSHEET
#
# Structure:
#
# HEADER
# DUN GROUP HEADER
# DM
# DM
# DM
# DUN TOTAL
# BLANK
# DUN GROUP HEADER
# DM
# DM
# DUN TOTAL
# ============================================================

def write_combined_dm_sheet(
    ws,
    final_df,
    age_groups,
    headers
):

    # ========================================================
    # MAIN HEADER
    # ========================================================

    write_main_header(
        ws,
        headers,
        row_number=1
    )

    current_row = 2

    # ========================================================
    # PREPARE DM SUMMARY
    # ========================================================

    dm_df = build_summary_df(
        final_df,
        [
            "_KOD DM",
            "_NAMA DM"
        ],
        age_groups,
        headers,
        kod_formatter=format_kod_dm
    )

    # ========================================================
    # ADD DUN SORTING KEY TO DM SUMMARY
    # ========================================================

    dm_df["_DUN_CODE"] = (
        dm_df["KOD"]
        .apply(get_dun_code_from_dm)
    )

    dm_df["_DUN_SORT"] = (
        dm_df["_DUN_CODE"]
        .apply(numeric_sort_key)
    )

    dm_df["_DM_SORT"] = (
        dm_df["KOD"]
        .apply(numeric_sort_key)
    )

    dm_df = (
        dm_df
        .sort_values(
            [
                "_DUN_SORT",
                "_DM_SORT"
            ],
            kind="stable"
        )
        .reset_index(drop=True)
    )

    # ========================================================
    # GROUP DM ROWS BY DUN
    # ========================================================

    for dun_code, dun_dm_df in dm_df.groupby(
        "_DUN_CODE",
        sort=False
    ):

        if not dun_code:
            continue

        # ----------------------------------------------------
        # Get DUN name from underlying records
        # ----------------------------------------------------

        dun_records = final_df[
            final_df["_KOD DUN"].apply(
                lambda x: format_kod_dun(x)
            ) == dun_code
        ]

        if dun_records.empty:
            continue

        dun_name_values = (
            dun_records["_NAMA DUN"]
            .dropna()
            .astype(str)
            .str.strip()
        )

        if dun_name_values.empty:
            dun_name = ""
        else:
            dun_name = (
                dun_name_values
                .iloc[0]
                .upper()
            )

        # ----------------------------------------------------
        # DUN GROUP HEADER
        # ----------------------------------------------------

        write_dun_group_header(
            ws,
            current_row,
            dun_code,
            dun_name,
            headers
        )

        current_row += 1

        # ----------------------------------------------------
        # DM ROWS
        # ----------------------------------------------------

        for _, dm_row in dun_dm_df.iterrows():

            values = [
                dm_row[header]
                for header in headers
            ]

            write_data_row(
                ws,
                current_row,
                values,
                headers,
                bold=False
            )

            current_row += 1

        # ----------------------------------------------------
        # DUN TOTAL
        #
        # IMPORTANT:
        # Calculated from underlying records belonging
        # to this DUN, NOT by summing/averaging percentages
        # from the DM summary rows.
        # ----------------------------------------------------

        dun_total = build_dun_total_row(
            dun_code,
            dun_name,
            dun_records,
            age_groups
        )

        write_dun_total_row(
            ws,
            current_row,
            dun_total,
            headers
        )

        current_row += 1

        # ----------------------------------------------------
        # BLANK ROW BETWEEN DUN GROUPS
        # ----------------------------------------------------

        current_row += 1

    # ========================================================
    # COLUMN WIDTHS / SETTINGS
    # ========================================================

    apply_column_widths(ws)
    apply_sheet_settings(ws)


# ============================================================
# PREPARE INPUT DATA
# ============================================================

def prepare_dataframe(
    df,
    age_groups
):

    df = df.copy()

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    # ========================================================
    # PARLIAMENT
    # ========================================================

    col_kod_parlimen = get_col(
        df,
        [
            "kod_parlimen",
            "KOD PARLIMEN",
            "KODPARLIMEN"
        ]
    )

    col_nama_parlimen = get_col(
        df,
        [
            "nama_parlimen",
            "NamaParlimen",
            "NAMA PARLIMEN"
        ]
    )

    # ========================================================
    # DUN
    # ========================================================

    col_kod_dun = get_col(
        df,
        [
            "kod_dun",
            "KOD DUN",
            "KODDUN"
        ]
    )

    col_nama_dun = get_col(
        df,
        [
            "nama_dun",
            "DUN",
            "NAMA DUN"
        ]
    )

    # ========================================================
    # DM
    # ========================================================

    col_dm = get_col(
        df,
        [
            "KOD DM",
            "kod_dm"
        ]
    )

    col_nama_dm = get_col(
        df,
        [
            "NamaDM",
            "nama_dm",
            "NAMA DM"
        ]
    )

    # ========================================================
    # OTHER COLUMNS
    # ========================================================

    col_jantina = get_col(
        df,
        [
            "JANTINA",
            "jantina"
        ]
    )

    col_bangsa = get_col(
        df,
        [
            "kaum",
            "BANGSA",
            "kategori_kaum"
        ]
    )

    col_umur = get_col(
        df,
        [
            "UMUR",
            "umur"
        ]
    )

    col_party = get_col(
        df,
        [
            "party",
            "PARTY"
        ]
    )

    col_sikap = get_col(
        df,
        [
            "CATATAN",
            "sikap"
        ]
    )

    col_no = get_col(
        df,
        [
            "NoPerkhidmatan",
            "noperkhidmatan"
        ]
    )

    col_pasangan = get_col(
        df,
        [
            "NoKPPasangan",
            "NoPerkhidmatanPasangan",
            "noperkhidmatanpasangan"
        ]
    )

    required = {
        "kod_parlimen": col_kod_parlimen,
        "nama_parlimen": col_nama_parlimen,
        "kod_dun": col_kod_dun,
        "nama_dun / DUN": col_nama_dun,
        "KOD DM": col_dm,
        "NamaDM": col_nama_dm,
        "JANTINA": col_jantina,
        "BANGSA": col_bangsa,
        "UMUR": col_umur,
        "NoPerkhidmatan": col_no,
        "NoKPPasangan": col_pasangan
    }

    missing = [
        key
        for key, value in required.items()
        if value is None
    ]

    if missing:

        raise ValueError(
            "Missing required columns: "
            + ", ".join(missing)
        )

    # ========================================================
    # HIERARCHY
    # ========================================================

    df["_KOD PARLIMEN"] = (
        df[col_kod_parlimen]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df["_NAMA PARLIMEN"] = (
        df[col_nama_parlimen]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df["_KOD DUN"] = (
        df[col_kod_dun]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df["_NAMA DUN"] = (
        df[col_nama_dun]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df["_KOD DM"] = (
        df[col_dm]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df["_NAMA DM"] = (
        df[col_nama_dm]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # ========================================================
    # DEMOGRAPHIC FIELDS
    # ========================================================

    df["_jantina"] = (
        df[col_jantina]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df["_race"] = (
        df[col_bangsa]
        .apply(normalise_race)
    )

    df["_age_group"] = (
        df[col_umur]
        .apply(
            lambda x:
            get_age_group(
                x,
                age_groups
            )
        )
    )

    # ========================================================
    # PARTY
    # ========================================================

    if col_party:

        df["_party"] = (
            df[col_party]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )

    else:

        df["_party"] = ""

    # ========================================================
    # SIKAP
    # ========================================================

    if col_sikap:

        df["_sikap"] = (
            df[col_sikap]
            .apply(normalise_sikap)
        )

    else:

        df["_sikap"] = ""

    # ========================================================
    # SERVICE NUMBERS
    # ========================================================

    df["_NoPerkhidmatan_clean"] = (
        df[col_no]
        .apply(clean_service_no)
    )

    df["_NoKPPasangan_clean"] = (
        df[col_pasangan]
        .apply(clean_service_no)
    )

    # ========================================================
    # CLASSIFICATION
    # ========================================================

    df["_awal_type"] = (
        df["_NoPerkhidmatan_clean"]
        .apply(classify_awal)
    )

    df["_Pasangan Polis"] = (
        df["_NoKPPasangan_clean"]
        .apply(
            lambda x:
            1 if is_polis(x) else 0
        )
    )

    df["_Pasangan Askar"] = (
        df["_NoKPPasangan_clean"]
        .apply(
            lambda x:
            1 if is_askar(x) else 0
        )
    )

    return df


# ============================================================
# GENERATE
# ============================================================

def generate_demografik(
    uploaded_files,
    age_groups
):

    logs = []

    all_data = []

    headers = build_headers(
        age_groups
    )

    # ========================================================
    # READ FILES
    # ========================================================

    for uploaded_file in uploaded_files:

        fname = uploaded_file.name

        try:

            df = pd.read_excel(
                uploaded_file,
                dtype=str
            )

            prepared = prepare_dataframe(
                df,
                age_groups
            )

            all_data.append(
                prepared
            )

            logs.append(
                f"Loaded {fname}: "
                f"{len(prepared):,} rows"
            )

        except Exception as e:

            logs.append(
                f"Error reading {fname}: "
                f"{e}"
            )

    if not all_data:

        raise ValueError(
            "No valid data loaded.\n"
            + "\n".join(logs)
        )

    final_df = pd.concat(
        all_data,
        ignore_index=True
    )

    # ========================================================
    # WORKBOOK
    # ========================================================

    wb = Workbook()

    default_ws = wb.active

    wb.remove(default_ws)

    # ========================================================
    # 1. PARLIMEN
    # ========================================================

    ws_parlimen = wb.create_sheet(
        title="PARLIMEN"
    )

    parlimen_df = build_summary_df(
        final_df,
        [
            "_KOD PARLIMEN",
            "_NAMA PARLIMEN"
        ],
        age_groups,
        headers,
        kod_formatter=format_kod_parlimen
    )

    parlimen_df = add_total_row(
        parlimen_df,
        headers
    )

    write_summary_sheet(
        ws_parlimen,
        parlimen_df,
        headers
    )

    logs.append(
        f"PARLIMEN worksheet: "
        f"{len(parlimen_df) - 1:,} Parliament(s)"
    )

    # ========================================================
    # 2. COMBINED DM WORKSHEET
    #
    # NO DUN WORKSHEET
    # ========================================================

    ws_dm = wb.create_sheet(
        title="DM"
    )

    write_combined_dm_sheet(
        ws_dm,
        final_df,
        age_groups,
        headers
    )

    # Count DUNs and DMs
    dun_count = (
        final_df[
            [
                "_KOD DUN",
                "_NAMA DUN"
            ]
        ]
        .drop_duplicates()
        .shape[0]
    )

    dm_count = (
        final_df[
            [
                "_KOD DM",
                "_NAMA DM"
            ]
        ]
        .drop_duplicates()
        .shape[0]
    )

    logs.append(
        f"DM worksheet: "
        f"{dm_count:,} DM(s) grouped under "
        f"{dun_count:,} DUN(s)"
    )

    logs.append(
        "DUN totals calculated directly "
        "from underlying records; percentages "
        "were recalculated from DUN counts."
    )

    logs.append(
        "Non-percentage statistical columns "
        "formatted with thousand separators."
    )

    logs.append(
        "DM data rows have no fill colour; "
        "only headers and total rows are coloured."
    )

    # ========================================================
    # SAVE
    # ========================================================

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    # ========================================================
    # OUTPUT NAME
    # ========================================================

    if len(uploaded_files) == 1:

        source_name = uploaded_files[0].name

        source_name = re.sub(
            r"\.(xlsx|xls)$",
            "",
            source_name,
            flags=re.IGNORECASE
        )

        out_name = (
            f"DEMOGRAFIK "
            f"{clean_filename(source_name)}.xlsx"
        )

    else:

        out_name = (
            "DEMOGRAFIK "
            f"({len(uploaded_files)} FILES).xlsx"
        )

    return (
        output.getvalue(),
        out_name,
        logs
    )
